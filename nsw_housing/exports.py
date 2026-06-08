"""
exports.py -- export NSW pipeline data to CSV + an Excel dashboard.

The raw `sales` table is transaction-level (often >1M rows), so it is NOT
exported wholesale. Instead we export a **suburb medians** aggregate plus the
smaller reference tables, and build a dashboard with a postcode-level yield sheet
(VG sale prices joined to DCJ rents).
"""
from __future__ import annotations

from .core import BASE_DIR, get_conn, get_logger

log     = get_logger("exports")
OUT_DIR = BASE_DIR / "exports"
OUT_DIR.mkdir(exist_ok=True)

SMALL_TABLES = ["rental_medians", "building_approvals", "lending_rates",
                "cash_rate", "capital_prices", "asx_announcements"]


def _suburb_medians():
    import pandas as pd
    conn = get_conn()
    df = pd.read_sql_query(
        "SELECT period, suburb, postcode, dwelling_type, price FROM sales "
        "WHERE price > 0 AND suburb != ''", conn)
    conn.close()
    if df.empty:
        return df
    df = df[df["period"].str.match(r"^\d{4}-Q\d$")]   # drop 'unknown'-dated sales
    g = (df.groupby(["period", "suburb", "postcode", "dwelling_type"])
           .agg(median_price=("price", "median"), num_sales=("price", "size"))
           .reset_index())
    g["median_price"] = g["median_price"].round(0)
    return g


def export_csvs() -> None:
    import csv
    conn = get_conn()
    for table in SMALL_TABLES:
        path = OUT_DIR / f"{table}.csv"
        try:
            cur = conn.execute(f"SELECT * FROM {table}")
            cols = [d[0] for d in cur.description]; rows = cur.fetchall()
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f); w.writerow(cols); w.writerows(rows)
            log.info(f"  CSV: {path.name} ({len(rows)} rows)")
        except Exception as e:
            log.warning(f"  CSV export failed for {table}: {e}")
    conn.close()
    sm = _suburb_medians()
    if not sm.empty:
        sm.to_csv(OUT_DIR / "suburb_medians.csv", index=False)
        log.info(f"  CSV: suburb_medians.csv ({len(sm):,} rows)")


def export_excel() -> None:
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        log.warning(f"Excel export needs pandas+openpyxl: {e}"); return

    path = OUT_DIR / "nsw_housing_dashboard.xlsx"
    sm = _suburb_medians()
    conn = get_conn()
    rent = pd.read_sql_query("SELECT * FROM rental_medians", conn)
    caps = pd.read_sql_query("SELECT * FROM capital_prices", conn)
    conn.close()

    # postcode-level yield: VG house median price vs DCJ postcode House rent
    yields = pd.DataFrame()
    if not sm.empty and not rent.empty:
        latest_p = sm["period"].max()
        sp = (sm[(sm.period == latest_p) & (sm.dwelling_type == "house")]
              .groupby("postcode")["median_price"].median().reset_index())
        rp = rent[(rent.region_type == "postcode") &
                  (rent.dwelling_type.str.contains("House", case=False, na=False))]
        if not rp.empty:
            rp = rp.groupby("region")["median_rent"].median().reset_index().rename(columns={"region": "postcode"})
            rp["postcode"] = rp["postcode"].astype(str).str.replace(r"\.0$", "", regex=True)
            sp["postcode"] = sp["postcode"].astype(str)
            yields = sp.merge(rp, on="postcode", how="inner")
            yields = yields[yields.median_price > 0]
            yields["gross_yield_pct"] = (yields["median_rent"] * 52 / yields["median_price"] * 100).round(2)
            yields = yields.sort_values("gross_yield_pct", ascending=False)

    wb = Workbook(); wb.remove(wb.active)
    HF = PatternFill("solid", fgColor="1F3864"); HFONT = Font(bold=True, color="FFFFFF", size=10)

    def write(name, df, maxrows=5000):
        ws = wb.create_sheet(title=name[:31])
        if df is None or df.empty:
            ws.cell(1, 1, "No data"); return
        df = df.head(maxrows)
        for ci, cn in enumerate(df.columns, 1):
            c = ws.cell(1, ci, cn); c.font = HFONT; c.fill = HF; c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, v in enumerate(row, 1):
                ws.cell(ri, ci, v)
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 32)
        ws.freeze_panes = "A2"

    latest = sm[sm.period == sm.period.max()] if not sm.empty else sm
    write("Suburb_Medians_Latest", latest.sort_values("median_price", ascending=False) if not sm.empty else sm)
    write("Postcode_Yields", yields)
    write("Rental_Medians", rent)
    write("Capital_Prices", caps)
    log.info(f"  Sheet 'Suburb_Medians_Latest': {0 if sm.empty else len(latest)} rows")
    log.info(f"  Sheet 'Postcode_Yields': {0 if yields.empty else len(yields)} rows")
    wb.save(path)
    log.info(f"Excel dashboard saved: {path}")


def run() -> None:
    log.info("Exporting CSVs...")
    export_csvs()
    log.info("Building Excel dashboard...")
    export_excel()
    log.info("Export complete.")
